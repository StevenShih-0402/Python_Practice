import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time

# 建立 User-Agent
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36 Edg/115.0.1901.200',
}

# 建立 Excel
wb = Workbook()
ws = wb.active
ws.title = "動畫瘋資訊"

header = ["動畫名稱","動畫年份","動畫集數","動畫觀看數", "動畫屬性", "動畫類型"]      # 匯入標題列
ws.append(header)
excel_row = 2

tags = ["動作", "冒險", "奇幻", "異世界", "魔法", "超能力", "科幻", "機甲", "校園", "喜劇", "戀愛", "青春", "勵志", "溫馨", "悠閒", "料理", "親情", "感人", "運動", "競技", "偶像", "音樂", "職場", "推理", "懸疑", "時間穿越", "歷史", "戰爭", "血腥暴力", "靈異神怪", "黑暗", "特攝", "BL", "GL"]
categorys = ["全部", "電影", "OVA", "雙語", "泡麵番", "真人演出"]
pages = [
    16, 6, 1 ,4 ,0, 4, 
    16, 4, 1, 3, 1, 1,
    13, 2, 1, 1, 1, 0,
    4, 1, 0, 1, 1, 0,
    6, 1, 1, 1, 1, 0,
    4, 1, 1, 1, 0, 1,
    11, 5, 1, 4, 0, 5,
    2, 1, 1, 0, 0, 0,
    13, 2, 1, 2, 1, 1,
    12, 1, 1, 2, 1, 1,
    9, 1, 1, 1, 1, 0, 
    7, 1, 1, 1, 1, 1, 
    5, 1, 1, 1, 0, 1,
    5, 1, 1, 1, 1, 1,
    3, 1, 1, 1, 1, 0,
    1, 0, 0, 1, 0, 0,
    2, 1, 1, 1, 1, 0,
    1, 1, 0, 1, 0, 0,
    4, 1, 1, 1, 0, 0,
    4, 1, 1, 1, 0, 0,
    3, 1, 1, 1, 0, 1,
    4, 1, 1, 1, 0, 1,
    3, 1, 1, 2, 1, 0,
    2, 1, 1, 1, 0, 0, 
    3, 1, 1, 1, 1, 0,
    1, 0, 0, 1, 0, 0,
    2, 1, 1, 1, 1, 0,
    3, 1, 1, 1, 0, 0,
    3, 1, 1, 1, 0, 0,
    4, 1, 0, 1, 1, 0,
    2, 1, 1, 1, 0, 0,
    6, 4, 1, 3, 0, 5,
    1, 1, 0, 0, 1, 0,
    1, 1, 0, 0, 0, 0
]

page_num = 0

# 爬取資料
for tag in tags:
    for category in categorys:
        if pages[page_num] == 0:
            page_num += 1
            continue
        
        for page in range(1, pages[page_num] + 1):
            r = requests.get('https://ani.gamer.com.tw/animeList.php?tags=' + tag + '&category=' + category + '&page=' + str(page), headers=headers)
            if r.status_code == 200:
                print(f'請求成功：{r.status_code}')

                soup = BeautifulSoup(r.text, 'html.parser')
                themelist_main = soup.find_all("a", class_="theme-list-main")

                for anime_item in themelist_main:
                    anime_name = anime_item.select_one(".theme-info-block > p").text.strip()
                    ws['A' + str(excel_row)] = anime_name
                    anime_time = anime_item.select_one(".theme-detail-info-block > p").text.strip()
                    ws['B' + str(excel_row)] = anime_time
                    anime_number = anime_item.select_one(".theme-detail-info-block > span").text.strip()
                    ws['C' + str(excel_row)] = anime_number
                    anime_viewer = anime_item.select_one(".theme-img-block > .show-view-number > p").text.strip()
                    ws['D' + str(excel_row)] = anime_viewer
                    
                    ws['E' + str(excel_row)] = tag
                    ws['F' + str(excel_row)] = category

                    excel_row += 1
            else:
                print(f'請求失敗：{r.status_code}')

        # 換下一個類型，中間間隔1.5秒，避免請求過快產生錯誤
        page_num += 1
        time.sleep(1.5)

wb.save("anigamer_data.xlsx")