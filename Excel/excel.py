# openpyxl 只支援2010以上的版本

from openpyxl import Workbook, load_workbook        #Workbook = Excel檔案

wb = load_workbook('D:\課程資料\下學期課表.xlsx')    #載入檔案
ws = wb.active                                      #選擇工作表(預設)
ws2 = wb['工作表1']

print(wb.sheetnames)                                #印出所有工作表名稱
wb.create_sheet('New WorkSheet.')

print(ws2['B3'].value)                               #特定Cell(格子)的值

ws['B15'].value = "testtest"                        #修改內容

wb.save('D:\課程資料\下學期課表.xlsx')               #儲存檔案，保留所有指令產生的更動