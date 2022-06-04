from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = [
    {
        'name':"機器小狗",
        'stars': 1,
        'atk': 2,
        'hel': 1,
        'ability': "聖盾術",
        'type': "Mech"
    },
    {
        'name':"克蘇恩的侍僧",
        'stars': 2,
        'atk': 2,
        'hel': 3,
        'ability': "嘲諷、重生",
        'type': "None"
    },
    {
        'name':"霹靂旋風",
        'stars': 3,
        'atk': 4,
        'hel': 1,
        'ability': "風怒、聖盾術",
        'type': "Elemental"
    },
    {
        'name':"長鬃草原獅",
        'stars': 4,
        'atk': 6,
        'hel': 5,
        'ability': "死亡之聲：召喚兩隻2/2土狼",
        'type': "Beast"
    },
    {
        'name':"吧嘎咕王",
        'stars': 5,
        'atk': 6,
        'hel': 3,
        'ability': "戰吼及死亡之聲：賦予你的其他魚人+2/+2",
        'type': "Mulroc"
    },
    {
        'name':"朋友的朋友",
        'stars': 6,
        'atk': 5,
        'hel': 6,
        'ability': "戰吼：發現一個夥伴",
        'type': "None"
    }
]

wb = Workbook()
ws = wb.active
ws.title = "手下"

header = ["名字","星數","攻擊力","血量","能力","種族"]      # 匯入標題列
ws.append(header)

for member in data:
    ws.append(list(member.values()))                      # 將字典的值轉成陣列，匯入資料表

for col in range(2,5):
    char = get_column_letter(col)
    ws[char + '8'] = f'=SUM({char + "2"}:{char + "7"})'    # 在第8行輸入公式

for col in range(1,7):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color="0033CCCC") # 將第一行的字體加粗並改成青綠色(色碼要去官網找)

wb.save("Member.xlsx")