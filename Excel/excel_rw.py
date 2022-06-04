from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("Creater.xlsx")
ws = wb.active

for row in range(1,5):
    for col in range(1,6):
        char = get_column_letter(col)       # 回傳數字對應的英文字母(1回傳A，2回傳B...以此類推)
        print(ws[char + str(row)].value)    # 讀取對應格子的值

ws.merge_cells("J2:K3")                     # 合併儲存格
ws.merge_cells('G1:I3')
ws.unmerge_cells('G1:I3')                   # 解除合併

ws.insert_rows(2)                           # 插入行
ws.insert_cols(3)                           # 插入列

ws.delete_rows(2)                           # 刪除行
ws.delete_cols(3)                           # 刪除列

ws.move_range("A1:E3", rows=1, cols=1)      # 移動特定範圍

wb.save("Creater.xlsx")